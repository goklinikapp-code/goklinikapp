from __future__ import annotations

import csv
import io
import secrets
import string
import unicodedata
from datetime import datetime
from pathlib import Path

from django.core.exceptions import ValidationError
from django.core.validators import validate_email
from django.db.models import Exists, OuterRef, Prefetch
from django.utils import timezone
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.pagination import PageNumberPagination
from rest_framework.response import Response

from apps.appointments.models import Appointment
from apps.chat.models import Message
from apps.pre_operatory.models import PreOperatory
from apps.users.models import GoKlinikUser

from .models import DoctorPatientAssignment, Patient
from .permissions import CanManagePatients
from .serializers import (
    AssignDoctorSerializer,
    PatientCreateUpdateSerializer,
    PatientDetailSerializer,
    PatientListSerializer,
)

ACTIVE_APPOINTMENT_STATUSES = [
    Appointment.StatusChoices.PENDING,
    Appointment.StatusChoices.CONFIRMED,
    Appointment.StatusChoices.IN_PROGRESS,
    Appointment.StatusChoices.RESCHEDULED,
]


HEADER_ALIASES = {
    "name": {"nome", "name", "full_name", "nome_completo"},
    "email": {"email", "e_mail", "mail"},
    "phone": {"telefone", "phone", "telefone_celular", "celular", "fone"},
}

TEMP_PASSWORD_LENGTH = 8
TEMP_PASSWORD_SPECIAL_CHARS = "!@#$%&"
TEMP_PASSWORD_LOWERCASE_CHARS = string.ascii_lowercase
TEMP_PASSWORD_UPPERCASE_CHARS = string.ascii_uppercase
TEMP_PASSWORD_DIGIT_CHARS = string.digits


def _normalize_header_value(value: object) -> str:
    normalized = unicodedata.normalize("NFKD", str(value or ""))
    normalized = normalized.encode("ascii", "ignore").decode("ascii")
    normalized = normalized.strip().lower().replace("-", "_").replace(" ", "_")
    while "__" in normalized:
        normalized = normalized.replace("__", "_")
    return normalized


def _resolve_column_indexes(header_row: list[object]) -> dict[str, int]:
    normalized_header = [_normalize_header_value(value) for value in header_row]
    indexes: dict[str, int] = {}
    for key, aliases in HEADER_ALIASES.items():
        for index, column_name in enumerate(normalized_header):
            if column_name in aliases:
                indexes[key] = index
                break
    missing_required = [field for field in ("name", "email", "phone") if field not in indexes]
    if missing_required:
        readable_names = {
            "name": "nome",
            "email": "email",
            "phone": "telefone",
        }
        missing_labels = ", ".join(readable_names[field] for field in missing_required)
        raise ValueError(f"Colunas obrigatórias ausentes: {missing_labels}.")
    return indexes


def _row_is_empty(row: list[object]) -> bool:
    return all(str(value or "").strip() == "" for value in row)


def _extract_row_value(row: list[object], index: int) -> str:
    if index >= len(row):
        return ""
    value = row[index]
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _split_name(full_name: str, *, fallback_email: str) -> tuple[str, str]:
    normalized = " ".join(full_name.strip().split())
    if not normalized:
        local_part = fallback_email.split("@", 1)[0].strip()
        normalized = local_part or "Paciente"
    parts = normalized.split(" ", 1)
    first_name = parts[0]
    last_name = parts[1] if len(parts) > 1 else ""
    return first_name, last_name


def _generate_temporary_password() -> str:
    chars = [
        secrets.choice(TEMP_PASSWORD_UPPERCASE_CHARS),
        secrets.choice(TEMP_PASSWORD_DIGIT_CHARS),
        secrets.choice(TEMP_PASSWORD_SPECIAL_CHARS),
    ]
    chars.extend(
        secrets.choice(TEMP_PASSWORD_LOWERCASE_CHARS)
        for _ in range(TEMP_PASSWORD_LENGTH - len(chars))
    )

    for index in range(len(chars) - 1, 0, -1):
        swap_index = secrets.randbelow(index + 1)
        chars[index], chars[swap_index] = chars[swap_index], chars[index]

    return "".join(chars)


def _read_csv_rows(uploaded_file) -> list[list[object]]:
    uploaded_file.seek(0)
    raw_bytes = uploaded_file.read()
    if isinstance(raw_bytes, str):
        text = raw_bytes
    else:
        for encoding in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                text = raw_bytes.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            text = raw_bytes.decode("utf-8", errors="ignore")
    return list(csv.reader(io.StringIO(text)))


def _read_xlsx_rows(uploaded_file) -> list[list[object]]:
    from openpyxl import load_workbook

    uploaded_file.seek(0)
    workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        return [list(row) for row in worksheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def _parse_import_file(uploaded_file) -> list[list[object]]:
    extension = Path(str(getattr(uploaded_file, "name", "") or "")).suffix.lower()
    content_type = (getattr(uploaded_file, "content_type", "") or "").lower()

    is_csv = extension == ".csv" or "csv" in content_type
    is_xlsx = extension == ".xlsx" or "spreadsheetml" in content_type

    if is_csv:
        return _read_csv_rows(uploaded_file)
    if is_xlsx:
        return _read_xlsx_rows(uploaded_file)

    raise ValueError("Formato de arquivo inválido. Envie CSV ou XLSX.")


class PatientPagination(PageNumberPagination):
    page_size = 25


class PatientViewSet(viewsets.ModelViewSet):
    permission_classes = [CanManagePatients]
    pagination_class = PatientPagination

    def get_queryset(self):
        user = self.request.user
        if not getattr(user, "is_authenticated", False) or not hasattr(user, "role"):
            return Patient.objects.none()

        professional_for_flags = None
        if user.role == GoKlinikUser.RoleChoices.SURGEON:
            professional_for_flags = user.id
        elif user.role in {
            GoKlinikUser.RoleChoices.CLINIC_MASTER,
            GoKlinikUser.RoleChoices.SECRETARY,
        }:
            professional_for_flags = (
                self.request.query_params.get("professional_id")
                or self.request.query_params.get("professional")
            )

        active_appointments = Appointment.objects.filter(
            patient_id=OuterRef("pk"),
            status__in=ACTIVE_APPOINTMENT_STATUSES,
        )
        completed_surgeries = Appointment.objects.filter(
            patient_id=OuterRef("pk"),
            appointment_type=Appointment.AppointmentTypeChoices.SURGERY,
            status=Appointment.StatusChoices.COMPLETED,
        )
        if professional_for_flags:
            active_appointments = active_appointments.filter(
                professional_id=professional_for_flags
            )
            completed_surgeries = completed_surgeries.filter(
                professional_id=professional_for_flags
            )

        queryset = Patient.objects.select_related(
            "tenant",
            "specialty",
            "doctor_assignment",
            "doctor_assignment__doctor",
            "doctor_assignment__assigned_by",
        ).prefetch_related(
            Prefetch(
                "pre_operatory_records",
                queryset=PreOperatory.objects.select_related("procedure").only(
                    "id",
                    "patient_id",
                    "status",
                    "updated_at",
                    "procedure_id",
                    "procedure__specialty_name",
                ),
            )
        ).annotate(
            has_active_appointment=Exists(active_appointments),
            has_completed_surgery=Exists(completed_surgeries),
        ).all()

        if user.role != GoKlinikUser.RoleChoices.SUPER_ADMIN:
            queryset = queryset.filter(tenant_id=user.tenant_id)
        if (
            user.role == GoKlinikUser.RoleChoices.SURGEON
            and self.action != "my_patients"
        ):
            queryset = queryset.filter(doctor_assignment__doctor_id=user.id)

        status_filter = self.request.query_params.get("status")
        app_status_filter = self.request.query_params.get("app_status")
        specialty_filter = self.request.query_params.get("specialty")
        pre_op_approved_raw = (
            self.request.query_params.get("pre_op_approved") or ""
        ).strip().lower()
        created_from = self.request.query_params.get("created_from")
        created_to = self.request.query_params.get("created_to")

        if status_filter:
            queryset = queryset.filter(status=status_filter)

        if app_status_filter == "installed":
            queryset = queryset.filter(app_installed_at__isnull=False)
        elif app_status_filter == "not_installed":
            queryset = queryset.filter(app_installed_at__isnull=True)

        if specialty_filter:
            queryset = queryset.filter(specialty_id=specialty_filter)

        if pre_op_approved_raw in {"1", "true", "yes"}:
            approved_pre_operatory = PreOperatory.objects.filter(
                patient_id=OuterRef("pk"),
                status=PreOperatory.StatusChoices.APPROVED,
            )
            queryset = queryset.annotate(
                has_approved_pre_operatory=Exists(approved_pre_operatory)
            ).filter(has_approved_pre_operatory=True)

        if created_from:
            queryset = queryset.filter(date_joined__date__gte=created_from)

        if created_to:
            queryset = queryset.filter(date_joined__date__lte=created_to)

        return queryset.order_by("-date_joined")

    def get_serializer_class(self):
        if self.action == "list":
            return PatientListSerializer
        if self.action in {"create", "update", "partial_update"}:
            return PatientCreateUpdateSerializer
        return PatientDetailSerializer

    def perform_create(self, serializer):
        serializer.save()

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)

        detail_serializer = PatientDetailSerializer(
            serializer.instance,
            context=self.get_serializer_context(),
        )
        headers = self.get_success_headers(detail_serializer.data)
        return Response(detail_serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop("partial", False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        detail_serializer = PatientDetailSerializer(
            serializer.instance,
            context=self.get_serializer_context(),
        )
        return Response(detail_serializer.data, status=status.HTTP_200_OK)

    def partial_update(self, request, *args, **kwargs):
        kwargs["partial"] = True
        return self.update(request, *args, **kwargs)

    @action(detail=False, methods=["get"], url_path="my-patients")
    def my_patients(self, request):
        user = request.user
        if user.role not in {
            GoKlinikUser.RoleChoices.SURGEON,
            GoKlinikUser.RoleChoices.NURSE,
            GoKlinikUser.RoleChoices.CLINIC_MASTER,
        }:
            return Response(status=status.HTTP_403_FORBIDDEN)

        queryset = self.get_queryset()
        if user.role == GoKlinikUser.RoleChoices.SURGEON:
            queryset = queryset.filter(doctor_assignment__doctor_id=user.id)
        elif user.role == GoKlinikUser.RoleChoices.NURSE:
            queryset = queryset.filter(doctor_assignment__isnull=False)
        else:
            professional_id = request.query_params.get("professional_id")
            if professional_id:
                queryset = queryset.filter(doctor_assignment__doctor_id=professional_id)
            else:
                queryset = queryset.filter(doctor_assignment__isnull=False)

        page = self.paginate_queryset(queryset)
        serializer = PatientListSerializer(
            page if page is not None else queryset,
            many=True,
            context=self.get_serializer_context(),
        )
        if page is not None:
            return self.get_paginated_response(serializer.data)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=True, methods=["get"], url_path="timeline")
    def timeline(self, request, pk=None):
        patient = self.get_object()

        items: list[dict] = []

        appointments = Appointment.objects.filter(patient=patient).order_by(
            "appointment_date", "appointment_time"
        )
        interactions = Message.objects.filter(room__patient=patient).order_by("created_at")

        for event in appointments:
            timestamp = datetime.combine(event.appointment_date, event.appointment_time)
            items.append(
                {
                    "timestamp": timestamp,
                    "type": "appointment",
                    "title": event.get_appointment_type_display(),
                    "description": event.notes,
                    "status": event.status,
                }
            )

        for event in interactions:
            items.append(
                {
                    "timestamp": event.created_at,
                    "type": "interaction",
                    "title": event.get_message_type_display(),
                    "description": event.content,
                    "status": "logged",
                }
            )

        items.sort(key=lambda item: item["timestamp"] or datetime.min)

        return Response(items, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"], url_path="assign-doctor")
    def assign_doctor(self, request, pk=None):
        user = request.user
        if user.role not in {
            GoKlinikUser.RoleChoices.CLINIC_MASTER,
            GoKlinikUser.RoleChoices.SECRETARY,
        }:
            return Response(status=status.HTTP_403_FORBIDDEN)

        patient = self.get_object()
        serializer = AssignDoctorSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        doctor = (
            GoKlinikUser.objects.filter(
                id=serializer.validated_data["doctor_id"],
                tenant_id=patient.tenant_id,
                role=GoKlinikUser.RoleChoices.SURGEON,
                is_active=True,
            )
            .order_by("id")
            .first()
        )
        if not doctor:
            return Response(
                {"doctor_id": ["Doctor not found for this tenant."]},
                status=status.HTTP_400_BAD_REQUEST,
            )

        DoctorPatientAssignment.objects.update_or_create(
            patient=patient,
            defaults={
                "doctor": doctor,
                "notes": serializer.validated_data.get("notes", "").strip(),
                "assigned_at": timezone.now(),
                "assigned_by": user,
            },
        )

        updated_patient = self.get_queryset().filter(pk=patient.pk).first()
        payload = PatientDetailSerializer(
            updated_patient or patient,
            context=self.get_serializer_context(),
        ).data
        return Response(payload, status=status.HTTP_200_OK)

    @action(
        detail=False,
        methods=["post"],
        url_path="import",
        parser_classes=[MultiPartParser, FormParser],
    )
    def import_patients(self, request):
        user = request.user
        if user.role not in {
            GoKlinikUser.RoleChoices.CLINIC_MASTER,
            GoKlinikUser.RoleChoices.SECRETARY,
        }:
            return Response(status=status.HTTP_403_FORBIDDEN)

        if not user.tenant_id:
            return Response(
                {"detail": "Usuário sem clínica vinculada."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        uploaded_file = request.FILES.get("file")
        if uploaded_file is None:
            return Response(
                {"file": ["Arquivo é obrigatório."]},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            rows = _parse_import_file(uploaded_file)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception:  # noqa: BLE001
            return Response(
                {"detail": "Não foi possível processar o arquivo enviado."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not rows:
            return Response(
                {"detail": "A planilha está vazia."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        header_row = rows[0]
        try:
            column_indexes = _resolve_column_indexes(header_row)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        total_rows = 0
        imported = 0
        duplicates = 0
        errors = 0
        error_details: list[str] = []

        for row_number, row in enumerate(rows[1:], start=2):
            if _row_is_empty(row):
                continue

            total_rows += 1
            email_value = _extract_row_value(row, column_indexes["email"]).lower()
            name_value = _extract_row_value(row, column_indexes["name"])
            phone_value = _extract_row_value(row, column_indexes["phone"])

            if not email_value:
                errors += 1
                if len(error_details) < 10:
                    error_details.append(f"Linha {row_number}: e-mail vazio.")
                continue

            try:
                validate_email(email_value)
            except ValidationError:
                errors += 1
                if len(error_details) < 10:
                    error_details.append(f"Linha {row_number}: e-mail inválido ({email_value}).")
                continue

            if GoKlinikUser.objects.filter(
                tenant_id=user.tenant_id,
                email__iexact=email_value,
            ).exists():
                duplicates += 1
                continue

            first_name, last_name = _split_name(name_value, fallback_email=email_value)
            temporary_password = _generate_temporary_password()

            try:
                Patient.objects.create_user(
                    tenant_id=user.tenant_id,
                    role=GoKlinikUser.RoleChoices.PATIENT,
                    first_name=first_name,
                    last_name=last_name,
                    email=email_value,
                    phone=phone_value,
                    status=Patient.StatusChoices.LEAD,
                    temp_password=temporary_password,
                    password=temporary_password,
                    is_active=True,
                )
            except Exception as exc:  # noqa: BLE001
                errors += 1
                if len(error_details) < 10:
                    error_details.append(f"Linha {row_number}: erro ao criar paciente ({exc}).")
                continue

            imported += 1

        return Response(
            {
                "total_rows": total_rows,
                "imported": imported,
                "duplicates": duplicates,
                "errors": errors,
                "error_details": error_details,
            },
            status=status.HTTP_200_OK,
        )
